"""Class-level statistics, grade curving, TeX/PDF compilation, review queue.

The largest of the merge-report internals: aggregates per-student summaries
into class artifacts (XML/MD/TeX/PDF), compiles per-student PDFs in parallel
via xelatex, and emits the review queue side-channel.
"""

from __future__ import annotations

import json
import os
import subprocess
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any

from xscore.marking.report_latex import (
    _class_report_to_tex, _student_report_to_tex,
)
from xscore.marking.report_markdown import _class_report_to_md
from xscore.marking.report_xml import class_report_to_xml
from xscore.shared.exam_paths import (
    artifact_class_grade_histogram_path,
    artifact_class_marks_xlsx_path,
    artifact_class_question_difficulty_path,
    artifact_class_report_combined_landscape_pdf_path,
    artifact_class_report_combined_portrait_2up_pdf_path,
    artifact_class_report_combined_portrait_pdf_path,
    artifact_class_report_md_path,
    artifact_class_report_pdf_2up_path,
    artifact_class_report_tex_path,
    artifact_class_report_xml_path,
    artifact_review_queue_json_path,
    artifact_review_queue_md_path,
    artifact_student_pdf_dir,
    artifact_student_pdfs_dir,
    artifact_student_report_pdf_portrait_2up_path,
    artifact_student_report_pdf_portrait_large_path,
    artifact_student_report_tex_landscape_path,
    artifact_student_report_tex_portrait_large_path,
    artifact_student_report_tex_portrait_path,
)
from eXercise.pdfjam_post import make_2up_landscape_pdf
from xscore.shared.terminal_ui import warn_line


# ---------------------------------------------------------------------------
# Env-var knobs (see default.env "Phase 8 — Reports" section)
# ---------------------------------------------------------------------------

def _grade_curve_target() -> int:
    """Read GRADE_CURVE_TARGET (default 80). Used as the env-var fallback
    when the natural-language prompt doesn't override the target."""
    raw = os.environ.get("GRADE_CURVE_TARGET", "80")
    try:
        return int(raw)
    except ValueError:
        warn_line(f"Invalid GRADE_CURVE_TARGET={raw!r} — using default 80")
        return 80


def _effective_curve_target(ctx: Any) -> int:
    """Resolve the curve target for *ctx*.

    Priority: ``ctx.instruction.curved_grade_override`` (if int) → env var
    ``GRADE_CURVE_TARGET`` (default 80).
    """
    instr = getattr(ctx, "instruction", None)
    if instr is not None:
        override = getattr(instr, "curved_grade_override", None)
        if override is not None:
            return int(override)
    return _grade_curve_target()


_TRUE_STRS  = {"true",  "1", "yes", "on"}
_FALSE_STRS = {"false", "0", "no",  "off"}


def _curved_grade_visible(ctx: Any) -> bool:
    """Resolve whether per-student PDFs include the curved % in their header.

    Priority: ``ctx.instruction.curved_grade_visible`` (if bool) → env var
    ``CURVED_GRADE_VISIBLE`` (default true). Unrecognised env values warn
    and fall back to True.
    """
    instr = getattr(ctx, "instruction", None)
    if instr is not None:
        override = getattr(instr, "curved_grade_visible", None)
        if override is not None:
            return bool(override)
    raw = os.environ.get("CURVED_GRADE_VISIBLE", "true").strip().lower()
    if raw in _TRUE_STRS:
        return True
    if raw in _FALSE_STRS:
        return False
    warn_line(f"Invalid CURVED_GRADE_VISIBLE={raw!r} — using default true")
    return True


def _xelatex_timeout() -> int:
    """Read XELATEX_TIMEOUT in seconds (default 60). Used by _compile_tex."""
    raw = os.environ.get("XELATEX_TIMEOUT", "60")
    try:
        return max(1, int(raw))
    except ValueError:
        warn_line(f"Invalid XELATEX_TIMEOUT={raw!r} — using default 60s")
        return 60


# ---------------------------------------------------------------------------
# Ranking + PDF/TeX glue
# ---------------------------------------------------------------------------

def _rank_students(students: list[dict]) -> list[dict]:
    """Return students sorted by percentage desc, each dict annotated with 'rank'.

    Ties share the same rank; the next rank skips (1, 2, 2, 4).
    Students with percentage=None sort last and receive rank=None.
    """
    sorted_s = sorted(
        students,
        key=lambda s: s["percentage"] if s["percentage"] is not None else -1,
        reverse=True,
    )
    rank = 1
    for i, s in enumerate(sorted_s):
        if i == 0:
            s["rank"] = rank if s["percentage"] is not None else None
        elif s["percentage"] is None:
            s["rank"] = None
        else:
            if s["percentage"] != sorted_s[i - 1]["percentage"]:
                rank = i + 1
            s["rank"] = rank
    return sorted_s


def _merge_pdfs(class_pdf: Path, students_dir: Path, output_pdf: Path, suffix: str) -> None:
    """Concatenate the class overview PDF with student PDFs matching ``*/*_<suffix>.pdf``."""
    student_pdfs = sorted(students_dir.glob(f"*/*_{suffix}.pdf"), key=lambda p: p.stem)

    try:
        from pikepdf import Pdf

        combined = Pdf.new()
        for pdf_path in [class_pdf, *student_pdfs]:
            if not pdf_path.exists():
                warn_line(f"PDF missing, skipping from combined report: {pdf_path.name}")
                continue
            with Pdf.open(pdf_path) as src:
                combined.pages.extend(src.pages)
        combined.save(output_pdf)
    except Exception as exc:  # noqa: BLE001
        warn_line(f"Could not create combined class report: {exc}")


def _compile_tex(tex_path: Path, output_dir: Path) -> None:
    """Compile .tex with xelatex. Warns on failure but does not raise."""
    try:
        result = subprocess.run(
            [
                "xelatex",
                "-interaction=nonstopmode",
                f"-output-directory={output_dir}",
                str(tex_path),
            ],
            capture_output=True,
            timeout=_xelatex_timeout(),
        )
        if result.returncode != 0:
            warn_line(
                f"xelatex returned {result.returncode} for {tex_path.name} — PDF may be missing"
            )
    except FileNotFoundError:
        warn_line("xelatex not found — PDF reports skipped (install TeX Live or MacTeX)")
    except subprocess.TimeoutExpired:
        warn_line(f"xelatex timed out for {tex_path.name}")
    except Exception as exc:  # noqa: BLE001
        warn_line(f"xelatex error for {tex_path.name}: {exc}")


# ---------------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------------

def _build_all_question_tables(
    questions: list,
    leaf_avgs: dict[str, float],
) -> tuple[dict[str, float], dict[str, int]]:
    """Return (all_avgs, all_max) for every question node including parents.

    Leaf averages come directly from leaf_avgs (keyed with _N suffixes for duplicates).
    Parent averages are the rounded sum of their direct children's averages (recursive).
    all_max is keyed with the same _N suffix convention using a seen counter.
    """
    from xscore.shared.models import flatten_questions

    def _subtree_avg(q) -> float | None:
        if not q.subquestions:
            return leaf_avgs.get(str(q.number or ""))
        parts = [_subtree_avg(c) for c in q.subquestions]
        valid = [p for p in parts if p is not None]
        return round(sum(valid), 1) if valid else None

    all_avgs: dict[str, float] = dict(leaf_avgs)
    all_max: dict[str, int] = {}
    seen: dict[str, int] = {}
    for q in flatten_questions(questions):
        num = str(q.number or "")
        if not num:
            continue
        seen[num] = seen.get(num, 0) + 1
        key = num if seen[num] == 1 else f"{num}_{seen[num]}"
        all_max[key] = int(q.marks or 0)
        if q.subquestions:
            avg = _subtree_avg(q)
            if avg is not None:
                all_avgs[key] = avg
    return all_avgs, all_max


def _apply_grade_curve(student_summaries: list[dict], target: int) -> None:
    """Compute curve offset (target − class_avg); add curved_pct to each summary in place."""
    known_pcts = [s["percentage"] for s in student_summaries if s["percentage"] is not None]
    class_avg = int(round(sum(known_pcts) / len(known_pcts))) if known_pcts else None
    curve_offset = (target - class_avg) if class_avg is not None else 0
    for s in student_summaries:
        s["curved_pct"] = (
            min(100, max(0, s["percentage"] + curve_offset))
            if s["percentage"] is not None else None
        )


# ---------------------------------------------------------------------------
# Pass 2 — per-student .tex files + parallel xelatex compile
# ---------------------------------------------------------------------------

def _pass2_write_tex(
    student_summaries: list[dict],
    full_reports: dict[str, dict],
    artifact_dir: Path,
    exam_name: str,
    workers: int,
    show_curved_grade: bool = True,
) -> None:
    """Write per-student .tex files (landscape + portrait + portrait-large), then compile all in parallel."""
    tex_paths: list[Path] = []
    for s in student_summaries:
        report = full_reports[s["name"]]
        report["curved_pct"] = s["curved_pct"]
        artifact_student_pdf_dir(artifact_dir, s["name"]).mkdir(parents=True, exist_ok=True)
        for orientation, path_fn, font_size in (
            ("landscape", artifact_student_report_tex_landscape_path,      10),
            ("portrait",  artifact_student_report_tex_portrait_path,       10),
            ("portrait",  artifact_student_report_tex_portrait_large_path, 12),
        ):
            tex_path = path_fn(artifact_dir, s["name"])
            tex_path.write_text(
                _student_report_to_tex(
                    report, exam_name=exam_name, orientation=orientation,
                    font_size=font_size, show_curved_grade=show_curved_grade,
                ),
                encoding="utf-8",
            )
            tex_paths.append(tex_path)

    with ThreadPoolExecutor(max_workers=workers) as ex:
        list(ex.map(lambda p: _compile_tex(p, p.parent), tex_paths))

    portrait_2up_jobs: list[tuple[Path, Path]] = []
    for s in student_summaries:
        p_in = artifact_student_report_pdf_portrait_large_path(artifact_dir, s["name"])
        p_out = artifact_student_report_pdf_portrait_2up_path(artifact_dir, s["name"])
        if p_in.is_file():
            portrait_2up_jobs.append((p_in, p_out))
    if portrait_2up_jobs:
        with ThreadPoolExecutor(max_workers=workers) as ex:
            list(ex.map(lambda j: make_2up_landscape_pdf(j[0], j[1]), portrait_2up_jobs))


# ---------------------------------------------------------------------------
# Class report assembly
# ---------------------------------------------------------------------------

def _write_class_marks_xlsx(
    *,
    class_report: dict,
    full_reports: dict[str, dict],
    scaffold_questions: list,
    out_path: Path,
) -> None:
    """Write a per-student × per-question marks grid as ``class_marks.xlsx``.

    One column per scaffold node (parents *and* leaves) in DFS order, plus
    Total / Raw % / Curved %. Parent columns roll up to the sum of their leaf
    descendants so a row's Total equals the sum of any complete level of the
    tree. Headers, max-marks row, and a class-average row at the bottom.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    from xscore.shared.models import flatten_questions, gradable_questions

    # Walk scaffold once. Column key matches per_question_max_marks keys
    # (same _N duplicate suffixing). For both parents and leaves,
    # gradable_questions([q]) returns the leaf set used for rollup.
    seen: dict[str, int] = {}
    columns: list[tuple[str, list[str]]] = []   # (column_key, leaf_keys_for_rollup)
    for q in flatten_questions(scaffold_questions):
        num = str(q.number or "")
        if not num:
            continue
        seen[num] = seen.get(num, 0) + 1
        key = num if seen[num] == 1 else f"{num}_{seen[num]}"
        leaf_keys = [str(c.number or "") for c in gradable_questions([q])]
        columns.append((key, leaf_keys))

    students = class_report["students"]
    max_marks = class_report["per_question_max_marks"]
    avgs = class_report["per_question_averages"]
    total_max = class_report["total_max_marks"]
    class_pct = class_report["class_average_pct"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Class marks"

    ws.append(["Student"] + [k for k, _ in columns] + ["Total", "Raw %", "Curved %"])
    ws.append(
        ["Max marks"]
        + [max_marks.get(k, "") for k, _ in columns]
        + [total_max, None, None]
    )

    def _sum_leaves(report: dict, leaf_keys: list[str]) -> float | None:
        by_num = {q["number"]: q.get("assigned_marks") for q in report.get("questions", [])}
        vals = [by_num.get(k) for k in leaf_keys]
        nums = [v for v in vals if v is not None]
        return sum(nums) if nums else None

    for s in students:
        report = full_reports.get(s["name"], {})
        row: list = [s["name"]]
        for _key, leaf_keys in columns:
            row.append(_sum_leaves(report, leaf_keys))
        row += [
            s.get("total_marks"),
            s["percentage"] / 100 if s.get("percentage") is not None else None,
            s["curved_pct"] / 100 if s.get("curved_pct") is not None else None,
        ]
        ws.append(row)

    # Class-average row — per_question_averages already covers parents
    # (subtree sums) so the row is internally consistent. Total is the
    # mean of known student totals to avoid double-counting parent rollups.
    known_totals = [s["total_marks"] for s in students if s.get("total_marks") is not None]
    avg_total = round(sum(known_totals) / len(known_totals), 1) if known_totals else None
    ws.append(
        ["Class average"]
        + [avgs.get(k, None) for k, _ in columns]
        + [avg_total, class_pct / 100 if class_pct is not None else None, None]
    )

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="EEEEEE")
    for cell in list(ws[1]) + list(ws[2]) + list(ws[ws.max_row]):
        cell.font = bold
        cell.fill = head_fill
    ws.freeze_panes = "B3"

    name_col_w = max(12, max((len(s["name"]) for s in students), default=10) + 2)
    ws.column_dimensions["A"].width = name_col_w
    for col_idx in range(2, 2 + len(columns)):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 6
    for offset, w in enumerate([10, 8, 8]):  # Total, Raw %, Curved %
        ws.column_dimensions[
            ws.cell(row=1, column=2 + len(columns) + offset).column_letter
        ].width = w

    raw_col = 2 + len(columns) + 1
    curve_col = 2 + len(columns) + 2
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=raw_col).number_format = "0%"
        ws.cell(row=r, column=curve_col).number_format = "0%"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _build_class_report(
    ctx: Any,
    student_summaries: list[dict],
    q_totals: dict[str, list[float]],
    exam_name: str,
) -> None:
    """Build and write class XML/MD/TeX/PDF. Runs after both passes."""
    total_max_marks = ctx.scaffold.total_marks
    leaf_avgs = {k: round(sum(v) / len(v), 1) for k, v in q_totals.items()}
    all_avgs, all_max = _build_all_question_tables(
        getattr(ctx.scaffold, "questions", []), leaf_avgs
    )
    known_pcts = [s["percentage"] for s in student_summaries if s["percentage"] is not None]
    class_avg = int(round(sum(known_pcts) / len(known_pcts))) if known_pcts else None
    per_question_pct: dict[str, int] = {
        qnum: int(round(avg / all_max[qnum] * 100))
        for qnum, avg in all_avgs.items()
        if all_max.get(qnum, 0) > 0
    }
    # Leaf-only percentages for the difficulty chart (parents would double-count).
    leaf_pct: dict[str, int] = {
        qnum: int(round(avg / all_max[qnum] * 100))
        for qnum, avg in leaf_avgs.items()
        if all_max.get(qnum, 0) > 0
    }
    # Charts are best-effort: if matplotlib isn't installed the LaTeX block
    # for the figure stays empty (template uses ``<% if histogram_path %>``).
    histogram_path: str | None = None
    difficulty_path: str | None = None
    try:
        from xscore.marking.class_charts import (
            render_grade_histogram, render_question_difficulty,
        )
        h = render_grade_histogram(
            student_summaries,
            artifact_class_grade_histogram_path(ctx.artifact_dir),
        )
        if h is not None:
            histogram_path = str(h)
        d = render_question_difficulty(
            leaf_pct, all_max,
            artifact_class_question_difficulty_path(ctx.artifact_dir),
        )
        if d is not None:
            difficulty_path = str(d)
    except ImportError:
        warn_line("matplotlib not installed — class report figures skipped")
    except Exception as exc:  # noqa: BLE001
        warn_line(f"class chart rendering failed: {type(exc).__name__}: {exc}")

    class_report = {
        "students": _rank_students(student_summaries),
        "per_question_averages": all_avgs,
        "per_question_max_marks": all_max,
        "per_question_pct_averages": per_question_pct,
        "class_average_pct": class_avg,
        "total_max_marks": total_max_marks,
        "histogram_path": histogram_path,
        "difficulty_path": difficulty_path,
    }
    artifact_class_report_xml_path(ctx.artifact_dir).write_text(
        class_report_to_xml(class_report), encoding="utf-8"
    )
    artifact_class_report_md_path(ctx.artifact_dir).write_text(
        _class_report_to_md(class_report), encoding="utf-8"
    )
    _write_class_marks_xlsx(
        class_report=class_report,
        full_reports=getattr(ctx, "full_reports", None) or {},
        scaffold_questions=getattr(ctx.scaffold, "questions", []),
        out_path=artifact_class_marks_xlsx_path(ctx.artifact_dir),
    )
    tex_path = artifact_class_report_tex_path(ctx.artifact_dir)
    tex_path.write_text(_class_report_to_tex(class_report, exam_name=exam_name), encoding="utf-8")
    _compile_tex(tex_path, tex_path.parent)
    _merge_pdfs(
        tex_path.with_suffix(".pdf"),
        artifact_student_pdfs_dir(ctx.artifact_dir),
        artifact_class_report_combined_landscape_pdf_path(ctx.artifact_dir),
        suffix="landscape",
    )
    _merge_pdfs(
        tex_path.with_suffix(".pdf"),
        artifact_student_pdfs_dir(ctx.artifact_dir),
        artifact_class_report_combined_portrait_pdf_path(ctx.artifact_dir),
        suffix="portrait",
    )

    class_pdf_path = tex_path.with_suffix(".pdf")
    class_2up_path = artifact_class_report_pdf_2up_path(ctx.artifact_dir)
    if class_pdf_path.is_file():
        make_2up_landscape_pdf(class_pdf_path, class_2up_path)
    if class_2up_path.is_file():
        _merge_pdfs(
            class_2up_path,
            artifact_student_pdfs_dir(ctx.artifact_dir),
            artifact_class_report_combined_portrait_2up_pdf_path(ctx.artifact_dir),
            suffix="portrait_2up",
        )


# ---------------------------------------------------------------------------
# Review queue (side channel)
# ---------------------------------------------------------------------------

def _write_review_queue(
    full_reports: dict[str, dict],
    artifact_dir: Path,
    collisions: list[dict] | None = None,
) -> int:
    """Emit a standalone list of marks the AI flagged as medium/low confidence.

    Returns the number of flagged confidence entries (also written to the
    JSON's ``"total"``). Cross-page mark collisions, if any, are appended to
    the same artifacts under ``"collisions"`` / ``"collisions_total"``.

    Pure side artifact: read by humans only, never by any pipeline step.
    Existing student/class reports and PDFs are unaffected by this code path.

    Each entry in the JSON file:
        {
          "student": ..., "question": ..., "confidence": "medium" | "low",
          "assigned_marks": ..., "max_marks": ...,
          "student_answer": ..., "correct_answer": ...,
          "explanation": ...    # truncated to ~200 chars for readability
        }

    Empty / missing confidence is treated as ``"high"`` and excluded.
    """
    entries: list[dict] = []
    for student_name in sorted(full_reports):
        report = full_reports[student_name]
        for q in report.get("questions") or []:
            conf = (q.get("confidence") or "").strip().lower()
            if conf in ("", "high"):
                continue
            if conf not in ("medium", "low"):
                # Unknown values still surface — they're an AI mistake worth seeing.
                pass
            explanation = str(q.get("explanation") or "")
            if len(explanation) > 200:
                explanation = explanation[:200].rstrip() + "…"
            entries.append({
                "student":        student_name,
                "question":       str(q.get("number", "")),
                "confidence":     conf,
                "assigned_marks": q.get("assigned_marks"),
                "max_marks":      q.get("max_marks"),
                "student_answer": str(q.get("student_answer") or ""),
                "correct_answer": str(q.get("correct_answer") or ""),
                "explanation":    explanation,
            })

    # Sort: low first, then medium; within each, by student then question.
    _conf_rank = {"low": 0, "medium": 1}
    entries.sort(key=lambda e: (_conf_rank.get(e["confidence"], 2), e["student"], e["question"]))

    coll = list(collisions or [])
    coll.sort(key=lambda c: (c.get("student", ""), str(c.get("question", "")), c.get("page", 0)))

    json_path = artifact_review_queue_json_path(artifact_dir)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps({
            "entries":          entries,
            "total":            len(entries),
            "collisions":       coll,
            "collisions_total": len(coll),
        }, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    # Markdown mirror — quick to skim.
    md_lines = [
        "# Review Queue",
        "",
        f"**{len(entries)} marks flagged for human review** "
        "(medium or low confidence; no impact on the marks already awarded).",
        "",
    ]
    if entries:
        md_lines += [
            "| Conf | Student | Q | Awarded | Max | Student Answer | Correct | Explanation |",
            "|------|---------|---|---------|-----|----------------|---------|-------------|",
        ]
        for e in entries:
            sa = (e["student_answer"] or "").replace("|", "/").replace("\n", " ")
            ca = str(e["correct_answer"] or "").replace("|", "/")
            ex = (e["explanation"] or "").replace("|", "/").replace("\n", " ")
            am = e["assigned_marks"]
            am_s = "?" if am is None else str(am)
            md_lines.append(
                f"| {e['confidence']} | {e['student']} | {e['question']} | {am_s} | "
                f"{e['max_marks']} | {sa} | {ca} | {ex} |"
            )
    else:
        md_lines.append("*No medium/low-confidence entries — the AI was confident on every question.*")

    if coll:
        md_lines += [
            "",
            "## Cross-page collisions",
            "",
            f"**{len(coll)} cross-page mark collision(s)** — same question scored on multiple pages.",
            "",
            "| Student | Q | Page | Earlier | Page | Winner |",
            "|---------|---|------|---------|------|--------|",
        ]
        for c in coll:
            md_lines.append(
                f"| {c['student']} | {c['question']} | {c['page']} | "
                f"{c['earlier_marks']} | {c['page_marks']} | {c['winner']} |"
            )

    artifact_review_queue_md_path(artifact_dir).write_text(
        "\n".join(md_lines) + "\n", encoding="utf-8"
    )

    return len(entries)
