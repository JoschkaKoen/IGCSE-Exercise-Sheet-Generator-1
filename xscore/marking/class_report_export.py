"""Side-channel artifact writers for the class report.

Two pure writers extracted from ``class_report.py``:

- :func:`_write_class_marks_xlsx` — per-student × per-question marks grid as
  ``class_marks.xlsx`` (called inline by ``_build_class_report``).
- :func:`_write_review_queue` — confidence audit of every marked question
  (sorted by ascending confidence) plus any cross-page collisions, as JSON
  + Markdown + plain text (called by ``merge_reports.build_review_queue``).

Both take their data as parameters and have no compile-time dependency on
``class_report``; ``class_report`` and ``merge_reports`` import them from
here.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from xscore.shared.exam_paths import (
    artifact_class_marks_xlsx_path,
    artifact_review_queue_json_path,
    artifact_review_queue_md_path,
    artifact_review_queue_txt_path,
)


def _write_class_marks_xlsx(
    *,
    class_report: dict,
    full_reports: dict[str, dict],
    scaffold_questions: list,
    out_path: Path,
    curve_target_pct: int | None = None,
) -> None:
    """Write a per-student × per-question marks grid as ``class_marks.xlsx``.

    Cells are written as **live formulas** so the teacher can edit a leaf
    mark in Excel and have the totals, percentages, and curved grades
    recalculate. Layout:

    - **Curve block** at rows 1–2: ``Curve target`` (editable, B1) and
      ``Curve offset`` (computed, B2). Per-student Curved % cells reference
      ``$B$2`` so changes to the target propagate.
    - **Table A** (every scaffold node, parents + leaves) and **Table B**
      (top-level only). Each: header / max-marks / per-student rows /
      class-average, with Total / Raw % / Curved % on the right.
    - Leaf cells hold plain numbers (the editable inputs). Parent cells in
      Table A are ``=SUM(<leaves under that parent>)``. Total / Raw % /
      Curved % and the class-average aggregates are formulas too.

    ``curve_target_pct`` is the integer 0–100 (typically read from
    ``class_stats.json``); falls back to env ``GRADE_CURVE_TARGET``, then 80
    — matches :func:`xscore.marking.class_report._grade_curve_target`.
    """
    import os

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from xscore.shared.models import flatten_questions, gradable_questions

    def _build_columns(qs: list) -> list[tuple[str, list[str], list[int], bool]]:
        # Each tuple: (display_key with _N suffix for duplicate numbers,
        # leaf_raw_nums = subtree leaf numbers for full-report rollup lookup,
        # leaf_col_indices = subtree leaf column indices in this table for
        # SUM formulas, is_scaffold_leaf flag). Identity-based id() lookup so
        # duplicate question numbers don't cross-link parent rollups.
        seen: dict[str, int] = {}
        keys: list[str] = []
        questions: list = []
        id_to_idx: dict[int, int] = {}
        for q in qs:
            num = str(q.number or "")
            if not num:
                continue
            seen[num] = seen.get(num, 0) + 1
            keys.append(num if seen[num] == 1 else f"{num}_{seen[num]}")
            questions.append(q)
            id_to_idx[id(q)] = len(keys) - 1
        cols: list[tuple[str, list[str], list[int], bool]] = []
        for key, q in zip(keys, questions):
            leaves = gradable_questions([q])
            is_scaffold_leaf = len(leaves) == 1 and leaves[0] is q
            leaf_raw_nums = [str(c.number or "") for c in leaves]
            leaf_col_indices = [id_to_idx[id(c)] for c in leaves if id(c) in id_to_idx]
            cols.append((key, leaf_raw_nums, leaf_col_indices, is_scaffold_leaf))
        return cols

    cols_all = _build_columns(flatten_questions(scaffold_questions))
    cols_top = _build_columns(scaffold_questions)

    # Workbook is the editable / lookup-by-name view, so order students
    # alphabetically here. Other class_report artifacts keep their rank order.
    students = sorted(
        class_report["students"],
        key=lambda s: str(s.get("name", "")).casefold(),
    )
    max_marks = class_report["per_question_max_marks"]
    avgs = class_report["per_question_averages"]
    total_max = class_report["total_max_marks"]

    # Resolve curve target: explicit arg → env → default 80.
    if curve_target_pct is None:
        env_val = os.environ.get("GRADE_CURVE_TARGET", "")
        try:
            curve_target_pct = int(env_val) if env_val else 80
        except ValueError:
            curve_target_pct = 80
    curve_target_pct = max(0, min(100, int(curve_target_pct)))

    def _sum_leaves(report: dict, leaf_raw_nums: list[str]) -> float | None:
        # Static rollup: sum of subtree-leaf marks looked up in the full
        # student report. Used for Table B (top-level rollups) and for
        # leaf cells (single-element list).
        by_num = {q.get("number"): q.get("assigned_marks") for q in report.get("questions", [])}
        nums = [v for v in (by_num.get(k) for k in leaf_raw_nums) if v is not None]
        return sum(nums) if nums else None

    wb = Workbook()
    ws = wb.active
    ws.title = "Class marks"

    bold = Font(bold=True)
    head_fill = PatternFill("solid", fgColor="EEEEEE")

    # ------------------------------------------------------------------
    # Curve block (rows 1–2). B2 is filled in after Table A rows are
    # known so the AVERAGE range is correct.
    # ------------------------------------------------------------------
    ws.cell(row=1, column=1, value="Curve target").font = bold
    ws.cell(row=1, column=2, value=curve_target_pct / 100).number_format = "0%"
    ws.cell(row=2, column=1, value="Curve offset").font = bold
    ws.cell(row=2, column=2).number_format = "0%"  # formula filled below

    next_row = 4  # leave row 3 blank as a visual separator

    def _append_table(
        columns: list[tuple[str, list[str], list[int], bool]],
        start_row: int,
    ) -> tuple[int, int, int]:
        """Write header, max-marks, per-student, class-average rows starting
        at ``start_row``. Returns ``(header_row, class_avg_row, total_col_idx)``;
        the % columns sit at ``total_col_idx + 1`` and ``+ 2``.
        """
        n_q = len(columns)
        total_col = 2 + n_q  # 1 = name, then n_q question cols
        raw_col = total_col + 1
        curved_col = total_col + 2

        header_row = start_row
        max_marks_row = start_row + 1
        student_first = start_row + 2
        student_last = student_first + len(students) - 1
        class_avg_row = student_last + 1

        # Per-question column letters (indexed 0..n_q-1) and the trailing trio.
        q_letter = [get_column_letter(2 + i) for i in range(n_q)]
        total_letter = get_column_letter(total_col)
        raw_letter = get_column_letter(raw_col)
        curved_letter = get_column_letter(curved_col)
        max_total_ref = f"${total_letter}${max_marks_row}"

        # Static columns hold the editable / authoritative numbers for this
        # table — Total sums these. In Table A: scaffold leaves only. In
        # Table B: every column (top-level rollup, leaves not present here).
        static_indices = [
            i for i, (_, _, leaf_col_indices, is_scaffold_leaf) in enumerate(columns)
            if is_scaffold_leaf or not leaf_col_indices
        ]

        # Header row.
        for c, val in enumerate(
            ["Student"] + [k for k, _, _, _ in columns] + ["Total", "Raw %", "Curved %"],
            start=1,
        ):
            ws.cell(row=header_row, column=c, value=val)

        # Max marks row — per-question max marks, then total_max literal.
        ws.cell(row=max_marks_row, column=1, value="Max marks")
        for i, (key, _, _, _) in enumerate(columns):
            ws.cell(row=max_marks_row, column=2 + i, value=max_marks.get(key, ""))
        ws.cell(row=max_marks_row, column=total_col, value=total_max)

        # Per-student rows.
        for s_idx, s in enumerate(students):
            r = student_first + s_idx
            report = full_reports.get(s["name"], {})
            ws.cell(row=r, column=1, value=s["name"])

            for i, (_, leaf_raw_nums, leaf_col_indices, is_scaffold_leaf) in enumerate(columns):
                cell = ws.cell(row=r, column=2 + i)
                if is_scaffold_leaf or not leaf_col_indices:
                    # Static value — leaf mark or top-level rollup from full
                    # report. Editable; Total formula sums these.
                    cell.value = _sum_leaves(report, leaf_raw_nums)
                else:
                    # Parent with all leaves present in this table — formula
                    # SUM over those leaf cells. Updates when a leaf is edited.
                    refs = ",".join(f"{q_letter[j]}{r}" for j in leaf_col_indices)
                    cell.value = f"=SUM({refs})"

            # Total: SUM of this row's static-value cells.
            if static_indices:
                refs = ",".join(f"{q_letter[j]}{r}" for j in static_indices)
                ws.cell(row=r, column=total_col, value=f"=SUM({refs})")
            else:
                ws.cell(row=r, column=total_col, value=0)

            # Raw % = Total / MaxTotal.
            ws.cell(
                row=r, column=raw_col,
                value=f"={total_letter}{r}/{max_total_ref}",
            ).number_format = "0%"
            # Curved % = MAX(0, MIN(1, Raw + offset)).
            ws.cell(
                row=r, column=curved_col,
                value=f"=MAX(0,MIN(1,{raw_letter}{r}+$B$2))",
            ).number_format = "0%"

        # Class-average row.
        ws.cell(row=class_avg_row, column=1, value="Class average")
        for i, (key, _, _, _) in enumerate(columns):
            ws.cell(row=class_avg_row, column=2 + i, value=avgs.get(key, None))
        ws.cell(
            row=class_avg_row, column=total_col,
            value=f"=AVERAGE({total_letter}{student_first}:{total_letter}{student_last})",
        )
        ws.cell(
            row=class_avg_row, column=raw_col,
            value=f"=AVERAGE({raw_letter}{student_first}:{raw_letter}{student_last})",
        ).number_format = "0%"
        ws.cell(
            row=class_avg_row, column=curved_col,
            value=f"=AVERAGE({curved_letter}{student_first}:{curved_letter}{student_last})",
        ).number_format = "0%"

        return header_row, class_avg_row, total_col

    # Top-level rollup table comes first — it's the at-a-glance view. The
    # detailed (per-subpart) breakdown follows under a heading.
    top_b, bot_b, total_b = _append_table(cols_top, next_row)

    # Fill in the curve-offset formula now that Table B's student row range
    # is known. Both tables share this offset (same students); anchoring on
    # the first-written table keeps cross-table references simple.
    n = len(students)
    if n > 0:
        raw_letter_b = get_column_letter(total_b + 1)
        student_first_b = top_b + 2
        student_last_b = student_first_b + n - 1
        ws.cell(
            row=2, column=2,
            value=f"=B1-AVERAGE({raw_letter_b}{student_first_b}:{raw_letter_b}{student_last_b})",
        )
    else:
        ws.cell(row=2, column=2, value=0)

    # Two blank rows + bold heading, then Table A (detailed breakdown).
    heading_row = bot_b + 3
    ws.cell(row=heading_row, column=1, value="Detailed breakdown").font = bold
    top_a, bot_a, total_a = _append_table(cols_all, heading_row + 2)

    # Bold + grey fill on header / max-marks / class-average rows.
    for top, bot in [(top_a, bot_a), (top_b, bot_b)]:
        for row_idx in (top, top + 1, bot):
            for cell in ws[row_idx]:
                cell.font = bold
                cell.fill = head_fill

    # Thin rule under each table's question-number header row, separating
    # the headers from the max-marks / student / class-average block.
    hdr_border = Border(bottom=Side(style="thin"))
    for top, total_col in [(top_b, total_b), (top_a, total_a)]:
        for c in range(1, total_col + 3):
            ws.cell(row=top, column=c).border = hdr_border

    ws.freeze_panes = "B6"

    # Left-align every written cell — curve block, both tables (incl. names,
    # Total / Raw % / Curved %), and the "Detailed breakdown" heading.
    left = Alignment(horizontal="left")
    for r, c in [(1, 1), (1, 2), (2, 1), (2, 2), (heading_row, 1)]:
        ws.cell(row=r, column=c).alignment = left
    for top, bot, total_col in [(top_a, bot_a, total_a), (top_b, bot_b, total_b)]:
        for r in range(top, bot + 1):
            for c in range(1, total_col + 3):
                ws.cell(row=r, column=c).alignment = left

    # Auto-fit column widths from cell content. Formulas (`=…`) skip
    # themselves — their column width comes from headers and static cells in
    # the same column, which is sufficient given the bounded values these
    # formulas produce (totals up to total_max, percentages up to 100%).
    widths_by_letter: dict[str, float] = {}

    def _track(cell: Any) -> None:
        v = cell.value
        if v is None or v == "":
            return
        if isinstance(v, str) and v.startswith("="):
            return
        if isinstance(v, (int, float)) and cell.number_format == "0%":
            text = f"{int(round(v * 100))}%"
        else:
            text = str(v)
        widths_by_letter[cell.column_letter] = max(
            widths_by_letter.get(cell.column_letter, 0), len(text)
        )

    for row in ws.iter_rows():
        for cell in row:
            _track(cell)

    for letter, w in widths_by_letter.items():
        ws.column_dimensions[letter].width = w + 1
    # Stored width includes ~5px cell padding; 4.04 = (3.33×7+5)/7 displays as 3.33 in Excel.
    ws.column_dimensions["B"].width = 4.04

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)



# ---------------------------------------------------------------------------
# Backwards-compat re-exports — review-queue helpers now live in
# :mod:`class_report_review_queue`.
# ---------------------------------------------------------------------------

from xscore.marking.class_report_review_queue import (  # noqa: E402, F401
    _marks_cell,
    _qnum_natural_key,
    _short_answer,
    _write_review_queue,
    format_review_entry_lines,
)
