"""Parse the student roster from any supported file format via Gemini."""

from __future__ import annotations

import csv
import io
import json
import os
import time
from pathlib import Path

def _load_prompt() -> str:
    """Return the rules-only system text for the read_student_list prompt.

    The .md file's SYSTEM section holds the schema, the header-skip rule,
    and the JSON output shape. Callers append the spreadsheet CSV (or attach
    the PDF separately) and send the combination as the user message.
    """
    from xscore.prompts.loader import load_prompt
    _, body = load_prompt("read_student_list", section="system")
    return body.rstrip()


_PROMPT = _load_prompt()


def _read_model_config() -> tuple[str, int | None, int | None]:
    from eXercise.ai_client import parse_model_spec
    raw = os.getenv("READ_STUDENT_LIST_MODEL", os.getenv("AI_DEFAULT_MODEL", "gemini-2.5-flash"))
    return parse_model_spec(raw)


_SHEET_KEYWORDS = ["student list", "student", "roster", "class list", "participants", "names"]


def _best_sheet(wb):
    """Return the worksheet whose name best matches a student-list keyword."""
    if len(wb.sheetnames) == 1:
        return wb.active
    for keyword in _SHEET_KEYWORDS:
        for name in wb.sheetnames:
            if keyword in name.lower():
                return wb[name]
    return wb.active or wb.worksheets[0]


def _spreadsheet_to_csv(path: Path) -> str:
    """Convert Excel or CSV to a plain CSV string."""
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required: pip install openpyxl>=3.1.0")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = _best_sheet(wb)
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in ws.iter_rows(values_only=True):
            writer.writerow([str(v) if v is not None else "" for v in row])
        wb.close()
        return buf.getvalue()
    elif ext == ".csv":
        return path.read_text(errors="replace")
    raise ValueError(f"Unsupported spreadsheet format: {path.suffix}")


def _parse_name_list(raw: str) -> list[str]:
    """Parse a list of names from either a bare JSON array or ``{"names": [...]}``."""
    data = json.loads(raw)
    if isinstance(data, list):
        return [str(x) for x in data]
    if isinstance(data, dict):
        for k in ("names", "students", "student_names", "items"):
            if k in data and isinstance(data[k], list):
                return [str(x) for x in data[k]]
    raise ValueError(f"Could not extract a name list from response: {raw[:200]!r}")


def read_student_list(folder: Path, artifact_dir: Path | None = None) -> list[str]:
    """Return student names from any student list file in *folder*.

    Supports .xlsx, .xls, .csv (text) and .pdf (Gemini Files API; rasterized PNG
    pages on OpenAI-compat). Uses ``READ_STUDENT_LIST_MODEL`` (or
    ``AI_DEFAULT_MODEL``) to choose the provider.

    Raises FileNotFoundError if no student list file is found.
    """
    candidates = [c for c in folder.glob("StudentList.*") if c.is_file()]
    if not candidates:
        for pat in ("*[Ss]tudent*", "*[Rr]oster*"):
            candidates = [c for c in folder.glob(pat) if c.is_file()]
            if candidates:
                break
    if not candidates:
        raise FileNotFoundError(f"No student list file found in {folder}")

    preferred = [f for f in candidates if "student" in f.name.lower()]
    target = preferred[0] if preferred else candidates[0]
    ext = target.suffix.lower()

    if ext not in (".xlsx", ".xls", ".csv", ".pdf"):
        raise ValueError(
            f"Unsupported student list format: {ext}. "
            "Supported: .xlsx, .xls, .csv, .pdf"
        )

    model_name, thinking_tokens, max_tokens = _read_model_config()

    from xscore.shared.terminal_ui import api_latency_line
    _save_prompt_path = None
    if artifact_dir is not None:
        from xscore.shared.exam_paths import artifact_student_list_prompt_path
        _save_prompt_path = artifact_student_list_prompt_path(artifact_dir)

    if model_name.startswith("gemini"):
        try:
            from google.genai import types as gai_types
        except ImportError:
            raise RuntimeError("google-genai not installed; run: pip install google-genai")
        from eXercise.ai_client import build_gemini_thinking_config, gemini_pdf_part, make_gemini_native_client
        client = make_gemini_native_client()
        if client is None:
            raise RuntimeError("GEMINI_API_KEY (or GOOGLE_API_KEY) not set")

        gen_config_kwargs: dict = {
            "system_instruction": _PROMPT,
            "max_output_tokens": max_tokens or 2048,
            "response_mime_type": "application/json",
            "response_schema": list[str],
        }
        if thinking_tokens is not None:
            gen_config_kwargs["thinking_config"] = build_gemini_thinking_config(thinking_tokens)
        gen_config = gai_types.GenerateContentConfig(**gen_config_kwargs)

        _prompt_user_text = ""
        if ext in (".xlsx", ".xls", ".csv"):
            csv_text = _spreadsheet_to_csv(target)
            _prompt_user_text = csv_text
            contents = [csv_text]
        else:  # .pdf
            from xscore.shared.terminal_ui import announce_ai_input  # noqa: PLC0415
            announce_ai_input(kind="PDF", note="Gemini, native bytes")
            _prompt_user_text = "(PDF attached)"
            contents = [
                gemini_pdf_part(client, target, label="student list"),
            ]

        if _save_prompt_path is not None:
            from xscore.shared.prompt_logger import attachment_part, save_prompt
            if ext in (".xlsx", ".xls", ".csv"):
                _audit_user: object = _prompt_user_text
            else:  # .pdf — mirror the Gemini contents (system + PDF)
                _audit_user = [
                    attachment_part(target.read_bytes(), "application/pdf"),
                ]
            save_prompt(
                _save_prompt_path, model=model_name,
                messages=[
                    {"role": "system", "content": _PROMPT},
                    {"role": "user", "content": _audit_user},
                ],
            )

        from eXercise.api_retry import retry_api_call  # noqa: PLC0415
        _t0 = time.perf_counter()
        response = retry_api_call(
            lambda: client.models.generate_content(
                model=model_name,
                contents=contents,
                config=gen_config,
            ),
            label="Student list",
        )
        api_latency_line(time.perf_counter() - _t0, label="student list")
        from eXercise.ai_client import split_gemini_response  # noqa: PLC0415
        raw, thinking_text = split_gemini_response(response)
    else:
        # OpenAI-compat path (Qwen, Grok, …): rasterize PDFs; send CSV as text.
        import base64 as _base64
        from eXercise.ai_client import (
            build_completion_kwargs,
            collect_streamed_response,
            make_ai_client,
            provider_for_model,
        )
        _result = make_ai_client(
            model_env="READ_STUDENT_LIST_MODEL",
            legacy_model_env="AI_DEFAULT_MODEL",
        )
        if _result is None:
            raise RuntimeError(
                f"READ_STUDENT_LIST_MODEL={model_name} requires the API key for "
                f"provider '{provider_for_model(model_name)}' in .env"
            )
        _oa_client, _, _provider, _, _ = _result
        _use_stream, _kw = build_completion_kwargs(
            _provider, thinking_tokens, max_tokens or 2048,
        )

        _prompt_user_text = ""
        if ext in (".xlsx", ".xls", ".csv"):
            csv_text = _spreadsheet_to_csv(target)
            _prompt_user_text = csv_text
            user_content = csv_text
        else:  # .pdf — rasterize at 200 DPI to keep request size sane
            import fitz as _fitz
            from xscore.shared.terminal_ui import announce_ai_input  # noqa: PLC0415
            announce_ai_input(kind="PNG", dpi=200, note="raster fallback")
            _prompt_user_text = f"[PDF: {target.name}]"
            with _fitz.open(str(target)) as _doc:
                _pages_b64 = [
                    _base64.b64encode(_doc[i].get_pixmap(dpi=200).tobytes("png")).decode()
                    for i in range(_doc.page_count)
                ]
            user_content = [
                {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{b64}"}}
                for b64 in _pages_b64
            ]

        if _save_prompt_path is not None:
            from xscore.shared.prompt_logger import save_prompt
            # user_content is plain text for csv/xlsx or list-of-parts (image_url) for pdf.
            save_prompt(
                _save_prompt_path, model=model_name,
                messages=[
                    {"role": "system", "content": _PROMPT},
                    {"role": "user", "content": user_content},
                ],
            )

        _msgs = [
            {"role": "system", "content": _PROMPT},
            {"role": "user", "content": user_content},
        ]
        _t0 = time.perf_counter()
        thinking_text = ""
        from eXercise.api_retry import retry_api_call  # noqa: PLC0415
        if _use_stream:
            def _do_stream() -> tuple[str, str]:
                _th: list[str] = []
                _stream = _oa_client.chat.completions.create(
                    model=model_name, messages=_msgs, stream=True, **_kw,
                )
                _raw = collect_streamed_response(_stream, thinking_out=_th)
                return _raw, "".join(_th)

            raw, thinking_text = retry_api_call(_do_stream, label="Student list (stream)")
        else:
            def _do_json() -> tuple[str, str]:
                _resp = _oa_client.chat.completions.create(
                    model=model_name, messages=_msgs,
                    response_format={"type": "json_object"}, **_kw,
                )
                return (
                    _resp.choices[0].message.content or "",
                    getattr(_resp.choices[0].message, "reasoning_content", "") or "",
                )

            def _do_plain() -> tuple[str, str]:
                _resp = _oa_client.chat.completions.create(
                    model=model_name, messages=_msgs, **_kw,
                )
                return (
                    _resp.choices[0].message.content or "",
                    getattr(_resp.choices[0].message, "reasoning_content", "") or "",
                )

            try:
                raw, thinking_text = retry_api_call(_do_json, label="Student list (json)")
            except Exception:
                # Provider may reject response_format=json_object — fall through to plain.
                raw, thinking_text = retry_api_call(_do_plain, label="Student list (plain)")
        api_latency_line(time.perf_counter() - _t0, label="student list")

    if _save_prompt_path is not None:
        from xscore.shared.prompt_logger import save_response  # noqa: PLC0415
        save_response(_save_prompt_path, raw, thinking=thinking_text)

    try:
        names = _parse_name_list(raw)
    except (json.JSONDecodeError, ValueError) as exc:
        raise RuntimeError(
            f"Student-list API returned non-JSON (model={model_name}): {exc}\n"
            f"  Raw response: {raw!r:.200}"
        ) from exc
    if _save_prompt_path is not None:
        try:
            from xscore.shared.prompt_logger import save_output_data
            save_output_data(
                _save_prompt_path, json.dumps({"names": names}, indent=2),
                ext="json",
            )
        except Exception:  # noqa: BLE001
            pass
    return names
