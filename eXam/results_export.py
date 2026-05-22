"""SQLite → XLSX export for a single test or all tests for a class."""

from __future__ import annotations

import json
from io import BytesIO

import openpyxl

from .db import connect


def export_test_xlsx(test_id: str) -> bytes:
    """One workbook: per-student rows × per-question columns + totals, plus a topic-rollup sheet."""
    with connect() as conn:
        test = conn.execute(
            "SELECT id, title, subject, class_label, question_ids FROM tests WHERE id=?",
            (test_id,),
        ).fetchone()
        if test is None:
            raise ValueError(f"no such test: {test_id}")
        qids = json.loads(test["question_ids"])
        # Pull the latest attempt per (student, question).
        rows = conn.execute(
            """
            SELECT a.student_id, s.name, s.class_label,
                   a.question_id, a.attempt_number, a.assigned_marks, a.max_marks,
                   a.submitted, a.reasoning, a.submitted_at,
                   a.hint_used, a.solution_revealed, a.example_used, a.kb_used
            FROM attempts a
            JOIN students s ON s.id = a.student_id
            WHERE a.test_id = ?
            ORDER BY s.class_label, s.name, a.question_id, a.attempt_number DESC
            """,
            (test_id,),
        ).fetchall()
    # Group: per (student) → per qid → latest attempt + attempt count
    by_student: dict[int, dict] = {}
    for r in rows:
        sid = r["student_id"]
        s = by_student.setdefault(
            sid,
            {"name": r["name"], "class_label": r["class_label"], "qs": {}, "attempts": {}},
        )
        qid = r["question_id"]
        s["attempts"][qid] = s["attempts"].get(qid, 0) + 1
        if qid not in s["qs"]:
            s["qs"][qid] = {
                "marks": float(r["assigned_marks"]),
                "max": float(r["max_marks"]),
                "submitted": r["submitted"],
                "reasoning": r["reasoning"],
            }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks"
    header = ["Name", "Class"] + [f"Q{i+1} ({qid.split('::')[-1]})" for i, qid in enumerate(qids)] + ["Total", "Max"]
    ws.append(header)
    for sid, s in by_student.items():
        row = [s["name"], s["class_label"] or ""]
        total = 0.0
        max_total = 0.0
        for qid in qids:
            cell = s["qs"].get(qid)
            if cell is None:
                row.append("")
            else:
                attempt_count = s["attempts"][qid]
                suffix = f" (×{attempt_count})" if attempt_count > 1 else ""
                row.append(f"{cell['marks']:.1f}/{cell['max']:.0f}{suffix}")
                total += cell["marks"]
                max_total += cell["max"]
        row.extend([f"{total:.1f}", f"{max_total:.0f}"])
        ws.append(row)

    # Per-question rollup (correctness rate, mean attempts).
    ws2 = wb.create_sheet("Per question")
    ws2.append(["Question", "Attempted by", "Correct rate", "Mean attempts"])
    for qid in qids:
        attempted = 0
        correct = 0
        att_total = 0
        for s in by_student.values():
            cell = s["qs"].get(qid)
            if cell is None:
                continue
            attempted += 1
            att_total += s["attempts"][qid]
            if cell["marks"] >= cell["max"] and cell["max"] > 0:
                correct += 1
        rate = (correct / attempted * 100) if attempted else 0.0
        mean_att = (att_total / attempted) if attempted else 0.0
        ws2.append([qid.split("::")[-1], attempted, f"{rate:.0f}%", f"{mean_att:.1f}"])

    # Test header on a third sheet for context.
    ws3 = wb.create_sheet("Test info")
    ws3.append(["test_id", test["id"]])
    ws3.append(["title", test["title"]])
    ws3.append(["subject", test["subject"]])
    ws3.append(["class_label", test["class_label"] or ""])
    ws3.append(["questions", len(qids)])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
